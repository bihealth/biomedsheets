# -*- coding: utf-8 -*-
"""Support code for writing, reading, and interpreting XLS-based sample
sheets
"""

import re

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation


class SetupColumnSizesMixin:

    def _setup_column_sizes(self, ws, space=2):
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0),
                                            len(cell.value)))
        for col, value in dims.items():
            ws.column_dimensions[col].width = dims[col] + space


class RareDisaseSheetCreator(SetupColumnSizesMixin):
    """Create rare disease XLSX sheet"""

    def run(self, output_path):
        """Create empty sheet for rare disease scheme to ``path``"""
        wb = Workbook()
        ws = wb.active
        self._setup_sheet(ws)
        wb.save(output_path)

    def _setup_sheet(self, ws):
        ws.title = 'Germline Sample Sheet'
        self._setup_header(ws)
        self._setup_validation(ws)
        self._add_example_rows(ws)
        self._setup_column_sizes(ws)
        ws.freeze_panes = ws['A2']

    def _setup_header(self, ws):
        ws.append((
            'DONOR ID',
            'FATHER ID',
            'MOTHER ID',
            'SEX',
            'AFFECTED',
            'BIO SAMPLE ID',
            'SAMPLE TYPE',
            'TEST SAMPLE ID',
            'LIBRARY ID',
            'LIBRARY TYPE',
            'KIT TYPE',
            'KIT VERSION',
        ))

    def _setup_validation(self, ws):
        """Setup data validation, creates list cells"""
        # Column: D/sex
        dv_sex = DataValidation(
            type='list', formula1='"male,female,unknown"', allow_blank=True)
        dv_sex.ranges.append('D2:D1000')
        dv_sex.error = 'Sex must be one of male, female, unknown'
        dv_sex.errorTitle = 'Invalid sex'
        ws.add_data_validation(dv_sex)
        # Column: E/affected state
        dv_affected = DataValidation(
            type='list', formula1='"yes,no,unknown"', allow_blank=True)
        dv_affected.ranges.append('E2:E1000')
        dv_affected.error = 'Affected must be one of yes, no, unknown'
        dv_affected.errorTitle = 'Invalid affected'
        ws.add_data_validation(dv_affected)
        # Column: G/sample type
        dv_source = DataValidation(
            type='list', formula1='"blood,saliva,unknown,other"',
            allow_blank=True)
        dv_source.ranges.append('G2:G1000')
        dv_source.error = (
            'Sample source must be one of blood, saliva, unknown, other')
        dv_source.errorTitle = 'Invalid sample source'
        ws.add_data_validation(dv_source)
        # Column: J/library type
        dv_library = DataValidation(
            type='list',
            formula1=('"Agilent_SureSelect_Human_All_Exon,'
                      'Illumina_TruSeq_DNA_PCR_Free_Library_Preparation_Kit,'
                      'Illumina_TruSeq_Stranded_mRNA_Library_Prep_Kit,'
                      'Illumina_TruSeq_Exome"'),
            allow_blank=True)
        dv_library.ranges.append('J2:J1000')
        dv_library.error = 'Kit type must be in the list'
        dv_library.errorTitle = 'Kit type'

    def _add_example_rows(self, ws):
        """Add example rows to sheet"""
        ws.append(('01_001', '01_002', '01_003', 'male', 'affected',
                   '01_001-S1', 'blood',
                   '01_001-S1-DNA1',
                   '01_001-S1-DNA1-WES1', 'WES',
                   'Agilent_SureSelect_Human_All_Exon', 'V6'))
        ws.append(('01_002', '', '', 'male', 'unaffected',
                   '01_001-S1', 'blood',
                   '01_001-S1-DNA1',
                   '01_001-S1-DNA1-WES1', 'WES',
                   'Agilent_SureSelect_Human_All_Exon', 'V6'))
        ws.append(('01_003', '', '', 'female', 'unaffected',
                   '01_001-S1', 'blood',
                   '01_001-S1-DNA1',
                   '01_001-S1-DNA1-WES1', 'WES',
                   'Agilent_SureSelect_Human_All_Exon', 'V6'))


class CancerMatchedSheetCreator(SetupColumnSizesMixin):
    """Create cancer matched tumor/normal sheet"""

    def run(self, output_path):
        """Create empty sheet for cancer matched study schema to ``path``"""
        wb = Workbook()
        ws = wb.active
        self._setup_sheet(ws)
        wb.save(output_path)

    def _setup_sheet(self, ws):
        ws.title = 'Cancer Matched Sample Sheet'
        self._setup_header(ws)
        self._setup_validation(ws)
        self._add_example_rows(ws)
        self._setup_column_sizes(ws)
        ws.freeze_panes = ws['A2']

    def _setup_header(self, ws):
        ws.append((
            'DONOR ID',
            'SEX',
            'BIO SAMPLE ID',
            'ICD-10 ENTRY (2016)',
            'IS TUMOR',
            'TNM STAGE',
            'PRESERVATION',
            'TEST SAMPLE ID',
            'EXTRACTION',
            'LIBRARY ID',
            'LIBRARY TYPE',
            'KIT TYPE',
            'KIT VERSION'
        ))

    def _setup_validation(self, ws):
        """Setup data validation, creates list cells"""
        # Column: B/sex
        dv_sex = DataValidation(
            type='list', formula1='"male,female,unknown"', allow_blank=True)
        dv_sex.ranges.append('B2:B1000')
        dv_sex.error = 'Sex must be one of male, female, unknown'
        dv_sex.errorTitle = 'Invalid sex'
        ws.add_data_validation(dv_sex)
        # Column: E/is cancer
        dv_tumor = DataValidation(
            type='list', formula1='"yes,no"', allow_blank=True)
        dv_tumor.ranges.append('E2:E1000')
        dv_tumor.error = 'Is tumor? must be one of yes, no, unknown'
        dv_tumor.errorTitle = 'Invalid flag'
        # Column: G/preservation
        dv_preservation = DataValidation(
            type='list', formula1='"FFPE,fresh-frozen,other"',
            allow_blank=True)
        dv_preservation.ranges.append('G2:G1000')
        dv_preservation.error = (
            'Preservation must be one of FFPE, fresh-frozen, other')
        dv_preservation.errorTitle = 'Invalid preservation'
        ws.add_data_validation(dv_preservation)
        # Column: I/extraction
        dv_extraction = DataValidation(
            type='list', formula1='"RNA,DNA,other"', allow_blank=True)
        dv_extraction.ranges.append('I2:I1000')
        dv_extraction.error = 'Preservation must be one of RNA, DNA, other'
        dv_extraction.errorTitle = 'Invalid extraction'
        ws.add_data_validation(dv_extraction)
        # Column: K/library type
        dv_library = DataValidation(
            type='list',
            formula1='"WES,Panel-seq,WGS,mRNA-seq,tRNA-seq,other"',
            allow_blank=True)
        dv_library.ranges.append('K2:K1000')
        dv_library.error = 'Library type must be in the list'
        dv_library.errorTitle = 'Invalid library type'
        ws.add_data_validation(dv_library)
        # Column: L/kit type
        dv_library = DataValidation(
            type='list',
            formula1=('"Agilent_SureSelect_Human_All_Exon,'
                      'Illumina_TruSeq_DNA_PCR_Free_Library_Preparation_Kit,'
                      'Illumina_TruSeq_Stranded_mRNA_Library_Prep_Kit,'
                      'Illumina_TruSeq_Exome"'),
            allow_blank=True)
        dv_library.ranges.append('L2:L1000')
        dv_library.error = 'Kit type must be in the list'
        dv_library.errorTitle = 'Kit type'
        ws.add_data_validation(dv_library)

    def _add_example_rows(self, ws):
        """Add example rows to sheet"""
        ws.append(('P001', 'male', 'P001-S1', '', 'no', '',
                   'other', 'P001-S1-DNA1', 'DNA',
                   'P001-S1-DNA1-WES1', 'WES',
                   'Agilent_SureSelect_Human_All_Exon', 'V6'))
        ws.append(('P001', 'male', 'P001-T1', 'C18.9', 'yes', 'T1 S1 M0',
                   'fresh-frozen', 'P001-T1-DNA1', 'DNA',
                   'P001-T1-DNA1-WES1', 'WES',
                   'Agilent_SureSelect_Human_All_Exon', 'V6'))
        ws.append(('P001', 'male', 'P001-T1', 'C18.9', 'yes', 'T1 S1 M0',
                   'fresh-frozen', 'P001-T1-RNA1', 'RNA',
                   'P001-T1-RNA1-RNAseq1', 'mRNA-seq',
                   'Illumina_TruSeq_Stranded_mRNA_Library_Prep_Kit', '1'))


class GenericExperimentSheetCreator(SetupColumnSizesMixin):

    def run(self, output_path):
        """Create empty sheet for generic experiments to ``path``"""
        wb = Workbook()
        ws = wb.active
        self._setup_sheet(ws)
        wb.save(output_path)

    def _setup_sheet(self, ws):
        ws.title = 'Generic Experiment Sample Sheet'
        self._setup_header(ws)
        self._setup_validation(ws)
        self._add_example_rows(ws)
        self._setup_column_sizes(ws)
        ws.freeze_panes = ws['A2']

    def _setup_header(self, ws):
        ws.append((
            'BIO ENTITY ID',
            'BIO SAMPLE ID',
            'TEST SAMPLE ID',
            'LIBRARY ID',
            'LIBRARY TYPE',
            'KIT TYPE',
            'KIT VERSION',
        ))

    def _setup_validation(self, ws):
        """Setup data validation, creates list cells"""
        # Column: E/library type
        dv_library = DataValidation(
            type='list',
            formula1='"WES,Panel-seq,WGS,mRNA-seq,tRNA-seq,other"',
            allow_blank=True)
        dv_library.ranges.append('E2:E1000')
        dv_library.error = 'Library type must be in the list'
        dv_library.errorTitle = 'Invalid library type'
        ws.add_data_validation(dv_library)
        # Column: F/kit type
        dv_kit = DataValidation(
            type='list',
            formula1=('"Agilent_SureSelect_Human_All_Exon,'
                      'Illumina_TruSeq_DNA_PCR_Free_Library_Preparation_Kit,'
                      'Illumina_TruSeq_Stranded_mRNA_Library_Prep_Kit,'
                      'Illumina_TruSeq_Exome"'),
            allow_blank=True)
        dv_kit.ranges.append('F2:F1000')
        dv_kit.error = 'Kit type must be in the list'
        dv_kit.errorTitle = 'Kit type'
        ws.add_data_validation(dv_kit)

    def _add_example_rows(self, ws):
        """Add example rows to sheet"""
        ws.append(('X001', 'X001-S1', 'X001-S1-RNA1', 'X001-S1-RNA1-RNAseq1',
                   'mRNA-seq',
                   'Illumina_TruSeq_Stranded_mRNA_Library_Prep_Kit', '1'))
        ws.append(('X002', 'X002-S1', 'X002-S1-RNA1', 'X001-S1-RNA1-RNAseq1',
                   'mRNA-seq',
                   'Illumina_TruSeq_Stranded_mRNA_Library_Prep_Kit', '1'))
        ws.append(('X003', 'X003-S1', 'X003-S1-RNA1', 'X001-S1-RNA1-RNAseq1',
                   'mRNA-seq',
                   'Illumina_TruSeq_Stranded_mRNA_Library_Prep_Kit', '1'))
        ws.append(('X004', 'X004-S1', 'X004-S1-RNA1', 'X001-S1-RNA1-RNAseq1',
                   'mRNA-seq',
                   'Illumina_TruSeq_Stranded_mRNA_Library_Prep_Kit', '1'))


class SheetIOBase:
    """Base class for XLSX reader and writer"""

    def __init__(self, sheet):
        #: models.Sheet to write
        self.sheet = sheet
        #: Shortcut to sheet's JSON data
        self.json_data = self.sheet.json_data
        #: Shortcut to extraInfoDefs dict of JSON data
        self.extra_info_defs = self.json_data.get('extraInfoDefs', {})

    def _bio_entity_headers(self):
        """Return headers for bio entities"""
        result = ['pk', 'secondary_id']
        result += list(self.extra_info_defs.get('bioEntity', {}).keys())
        return result

    def _bio_sample_headers(self):
        """Return headers for bio samples"""
        result = ['pk', 'secondary_id']
        result += list(self.extra_info_defs.get('bioSample', {}).keys())
        return result

    def _test_sample_headers(self):
        """Return headers for test samples"""
        result = ['pk', 'secondary_id']
        result += list(self.extra_info_defs.get('testSample', {}).keys())
        return result

    def _ngs_library_headers(self):
        """There must be at least one 'ngsLibraries' entry exist or
        'ngsLibrary' is set in the 'extraInfoDefs' section for any header
        output
        """
        if self.extra_info_defs.get('ngsLibrary') is None:
            return []
        else:
            result = ['pk', 'secondary_id']
            result += list(self.extra_info_defs.get('ngsLibrary', {}).keys())
            return result

    def _ms_protein_pool_headers(self):
        """There must be at least one 'msProteinPools' entry exist or
        'msProteinPool' is set in the 'extraInfoDefs' section for any header
        output
        """
        if self.extra_info_defs.get('msProteinPool') is None:
            return []
        else:
            result = ['pk', 'secondary_id']
            result += list(self.extra_info_defs['msProteinPool'].keys())
            return result


class SheetWriter(SheetIOBase, SetupColumnSizesMixin):
    """Writing of models.Sheet objects to XLSX"""

    def write(self, output_path):
        """Write out sheet to XLSX file"""
        wb = Workbook()
        ws = wb.active
        self._setup_sheet(ws)
        wb.save(output_path)

    def _setup_sheet(self, ws):
        """Setup openpyxl Worksheet ``ws``"""
        self._add_header(ws)
        self._setup_validation(ws)
        self._add_rows(ws)
        self._setup_column_sizes(ws)
        ws.freeze_panes = ws['A2']

    def _add_header(self, ws):
        """Adds header to the worksheet ``ws``"""

        def fn(s):
            """Local helper function for camelCase -> CAMEL CASE"""
            s = re.sub(r'([a-z])([A-Z])', r'\1_\2', s)
            return re.sub('[^a-zA-Z0-9]', ' ', s).upper()

        header = []
        header += ['DONOR ' + fn(s) for s in self._bio_entity_headers()]
        header += ['BIO SAMPLE ' + fn(s) for s in self._bio_sample_headers()]
        header += ['TEST SAMPLE ' + fn(s) for s in self._test_sample_headers()]
        header += ['NGS LIB ' + fn(s) for s in self._ngs_library_headers()]
        header += [
            'MS POOL ' + fn(s) for s in self._ms_protein_pool_headers()]
        ws.append(header)

    def _setup_validation(self, ws):
        """Setup validation for the sheet (validation for being in a list
        triggers dropdown lists)
        """
        pass

    def _add_rows(self, ws):
        """Add rows for all data in ``self.sheet``"""
        for bio_entity in self.sheet.bio_entities.values():
            for row in self._bio_entity_rows(bio_entity):
                ws.append(list(map(str, row)))

    def _bio_entity_rows(self, bio_entity):
        keys = self.extra_info_defs.get('bioEntity', {}).keys()
        cells = [bio_entity.pk, bio_entity.secondary_id]
        cells += [bio_entity.extra_infos.get(key, '') for key in keys]
        for bio_sample in bio_entity.bio_samples.values():
            yield from self._bio_sample_rows(bio_sample, cells)

    def _bio_sample_rows(self, bio_sample, base):
        keys = self.extra_info_defs.get('bioSample', {}).keys()
        cells = [bio_sample.pk, bio_sample.secondary_id]
        cells += [bio_sample.extra_infos.get(key, '') for key in keys]
        for test_sample in bio_sample.test_samples.values():
            yield from self._test_sample_rows(
                test_sample, base + cells)

    def _test_sample_rows(self, test_sample, base):
        keys = self.extra_info_defs.get('testSample', {}).keys()
        cells = [test_sample.pk, test_sample.secondary_id]
        cells += [test_sample.extra_infos.get(key, '') for key in keys]
        for ngs_library in test_sample.ngs_libraries.values():
            yield from self._ngs_library_rows(
                ngs_library, base + cells)
        for ms_protein_pool in test_sample.ms_protein_pools.values():
            base += [''] * len(self._ngs_library_headers())  # add spacers
            yield from self._ms_protein_pool_rows(
                ms_protein_pool, base + cells)

    def _ngs_library_rows(self, ngs_library, base):
        keys = self.extra_info_defs.get('ngsLibrary', {}).keys()
        cells = [ngs_library.pk, ngs_library.secondary_id]
        cells += [ngs_library.extra_infos.get(key, '') for key in keys]
        yield base + cells

    def _ms_protein_pool_rows(self, ms_protein_pool, base):
        keys = self.extra_info_defs.get('msProteinPool', {}).keys()
        cells = [ms_protein_pool.pk, ms_protein_pool.secondary_id]
        cells += [ms_protein_pool.extra_infos.get(key, '') for key in keys]
        yield base + cells


class SheetLoader(SheetIOBase):
    """Code for reading XLSX sheets

    For this to work, loader has to be initialized with a JSON sheet that
    carries all of the information.
    """

    def __init__(self, sheet):
        #: ``models.Sheet`` to use for data type information etc
        self.sheet = sheet

    def load(self, xlsx_path):
        """Load XLSX from file at ``xlsx_path`` and return models.Sheet
        instance
        """
        # Load XLSX from the given path
        wb = load_workbook(xlsx_path, read_only=True)
        # Match columns in XLSX and JSON to each other
        col_infos = self._get_col_infos(wb.active)
        # Load the entities from XLSX file
        return self._load_sheet(wb.active, col_infos)

    def _get_col_infos(self, ws):
        """Match columns from Worksheet ``ws`` to JSON properties"""
        raise NotImplementedError('Implement me!')

    def _load_sheet(self, ws, col_infos):
        """Load the information from the Worksheet ``ws``"""
        raise NotImplementedError('Implement me!')
